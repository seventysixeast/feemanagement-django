from django.contrib import admin
from .models import (
    transport, tuition_fees_defaulter,admission_report, final_fees_report, transport_defaulter, cheque_deposit, collection_report,activity_fees_defaulter,expected_fee_register
)

from app.models import (
    student_master, student_fee, student_class, specialfee_master,
    payment_schedule_master, latefee_master, fees_master, expense,
    concession_master, bus_master, busfees_master, account_head,generate_mobile_number_list
)

from import_export import resources
from import_export.admin import ExportMixin
from import_export.fields import Field


from django.db.models import OuterRef, Subquery
from django.urls import path
from django.http import JsonResponse
from django.shortcuts import render

# from .models import TuitionFeesDefaulter, student_master, student_classes
from datetime import date
from django.db.models.functions import Concat

from django.db.models.expressions import RawSQL

# from django.db.models import F, Value, CharField, ExpressionWrapper, Q
from django.db.models.functions import Coalesce
from django.db.models import DecimalField

from django.db.models import Case, When, Value, F, Q, CharField, ExpressionWrapper, DecimalField

from django.utils.html import format_html
from django.utils import timezone

from django.db.models import Subquery, OuterRef
from django.db.models import Exists, OuterRef


from django.db.models import F, Q, Value, CharField
    
from django.db.models import CharField, Value as V

from django.db.models import Sum, Max,F
from datetime import datetime, timedelta
from django import forms
from django.template.response import TemplateResponse

import csv
from django.http import HttpResponse
from django.urls import path
from django.shortcuts import render

from django.db import connection
import pandas as pd
from io import BytesIO
from django.contrib import messages
from django.shortcuts import redirect
from MySQLdb.cursors import DictCursor

from django.db import connection
from django.utils.dateparse import parse_date
import json



from datetime import date
from django.utils.html import format_html
from django.db.models import Sum, F, Value, CharField, IntegerField
from django.db.models.functions import Concat, Coalesce
from collections import defaultdict


from django.urls import path
from django.db import connection
from django.http import HttpResponse
from django.contrib import admin, messages
from io import BytesIO
import pandas as pd
from datetime import datetime

from openpyxl import Workbook

from django.utils import timezone
from django.utils.translation import gettext_lazy as _
import json
from decimal import Decimal


from django.contrib import admin
from django.http import HttpResponse
import openpyxl


# from .forms import DefaultersReportForm


# from .models import transport, BusFeesMaster, BusMaster, StudentClasses

class BusRouteFilter(admin.SimpleListFilter):
    title = 'Bus Route'
    parameter_name = 'bus_route'

    def lookups(self, request, model_admin):
        return []

class DestinationFilter(admin.SimpleListFilter):
    title = 'Destination'
    parameter_name = 'destination'

    def lookups(self, request, model_admin):
        return []

class ClassFilter(admin.SimpleListFilter):
    title = 'Class'
    parameter_name = 'class_no'

    def lookups(self, request, model_admin):
        return []

# Define class choices
CLASS_CHOICES = [
    ('', 'Select the Class'),
    ('Play-way', 'Play-way'),
    ('Nursery', 'Nursery'),
    ('lkg', 'LKG'),
    ('ukg', 'UKG'),
    ('1', '1'),
    ('2', '2'),
    ('3', '3'),
    ('4', '4'),
    ('5', '5'),
    ('6', '6'),
    ('7', '7'),
    ('8', '8'),
    ('9', '9'),
    ('10', '10'),
    ('11', '11'),
    ('12', '12')
]

# def get_months_array(year):
#     start_date = date(year, 4, 1)
#     end_date = date(year + 1, 3, 31)
#     current_date = date.today()

#     months = []
#     time = start_date
#     if current_date < end_date:
#         end_date = current_date

#     while time <= end_date:
#         months.append(time.month)
#         next_month = (time.month % 12) + 1
#         next_year = time.year + (1 if next_month == 1 else 0)
#         time = date(next_year, next_month, 1)

#     return months

from datetime import datetime, timedelta

def get_months_array(year):
    # Define the start and end dates of the financial year
    datefrom = datetime(year, 4, 1)  # April 1st of the selected year
    dateto = datetime(year + 1, 3, 31)  # March 31st of the next year
    
    # Get current date
    currentdate = datetime.today()

    montharray = []

    # If the current date is before the financial year end
    if currentdate < dateto:
        # Adjust the end date to today's date
        dateto = currentdate

        # Generate months from datefrom to dateto
        cur_date = datefrom
        while cur_date <= dateto:
            # Append the month number without leading zero
            montharray.append(cur_date.month)
            # Move to the next month
            cur_date = (cur_date.replace(day=28) + timedelta(days=4)).replace(day=1)
        
        # Sort the array to ensure months are in ascending order
        montharray.sort()
    else:
        # If the current date is after the financial year end, return all months
        montharray = list(range(1, 13))  # Months 1 to 12

    return montharray


def calculate_unpaid_months(paid_months, months_array):
        # print('paid_months in calculate_unpaid_months', paid_months)
        # print('months_array in calculate_unpaid_months', months_array)
        if not paid_months:
            return ','.join(map(str, months_array))  # Return all months if no months are paid

        paid_months_list = list(map(int, paid_months.split(',')))
        unpaid_months = set(months_array) - set(paid_months_list)
        return ','.join(map(str, sorted(unpaid_months)))

class TransportMasterResource(resources.ModelResource):

    class Meta:
        model = transport
        fields = ('student_name', 'addmission_no', 'class_no', 'section', 'destination', 'route', 
        'father_name', 'phone_no', 'bus_driver', 'driver_phone', 'bus_conductor', 
        'conductor_phone', 'bus_attendant', 'attendant_phone')

    student_name = Field(attribute='student_name', column_name='Student Name')
    addmission_no = Field(attribute='addmission_no', column_name='Admission No')
    class_no = Field(attribute='class_no', column_name='Class')
    section = Field(attribute='section', column_name='Section')
    destination = Field(attribute='destination', column_name='Destination')

    route = Field(attribute='route', column_name='Route')
    father_name = Field(attribute='father_name', column_name='Father Name')
    phone_no = Field(attribute='phone_no', column_name='Phone No.')
    bus_driver = Field(attribute='bus_driver', column_name='Bus Driver')
    driver_phone = Field(attribute='driver_phone', column_name='Driver Phone')

    bus_conductor = Field(attribute='bus_conductor', column_name='Bus Conductor')
    conductor_phone = Field(attribute='conductor_phone', column_name='Conductor Phone')
    bus_attendant = Field(attribute='bus_attendant', column_name='Bus Attendant')
    attendant_phone = Field(attribute='attendant_phone', column_name='Attendant Phone')

    def dehydrate_class_no(self, obj):
        # Assuming 'bus_driver' is a related field on the model
        student_class_instance = student_class.objects.filter(student_id=obj.student_id).order_by('-started_on').first()
        return student_class_instance.class_no if student_class_instance else None
        # return obj.class_no if obj.class_no else ''

    
    def dehydrate_section(self, obj):
        student_class_instance = student_class.objects.filter(student_id=obj.student_id).order_by('-started_on').first()
        return student_class_instance.section if student_class_instance else None
    # get_section.short_description = 'Section'


    def dehydrate_route(self, obj):
        busfees_master_instance = busfees_master.objects.filter(bus_id=obj.bus_id).first()
        return busfees_master_instance.route if busfees_master_instance else None
        # return obj.route

    def dehydrate_destination(self, obj):
        busfees_master_instance = busfees_master.objects.filter(bus_id=obj.bus_id).first()
        return busfees_master_instance.destination if busfees_master_instance else None
        # return obj.destination


    def dehydrate_bus_driver(self, obj):
        busfees_master_instance = busfees_master.objects.filter(bus_id=obj.bus_id).first()
        bus_master_instance = bus_master.objects.filter(bus_route=busfees_master_instance.route).first() if busfees_master_instance else None
        return bus_master_instance.bus_driver if bus_master_instance else None
        # return obj.bus_driver
    
    def dehydrate_driver_phone(self, obj):
        busfees_master_instance = busfees_master.objects.filter(bus_id=obj.bus_id).first()
        bus_master_instance = bus_master.objects.filter(bus_route=busfees_master_instance.route).first() if busfees_master_instance else None
        return bus_master_instance.driver_phone if bus_master_instance else None

    def dehydrate_bus_conductor(self, obj):
        busfees_master_instance = busfees_master.objects.filter(bus_id=obj.bus_id).first()
        bus_master_instance = bus_master.objects.filter(bus_route=busfees_master_instance.route).first() if busfees_master_instance else None
        return bus_master_instance.bus_conductor if bus_master_instance else None
        # return obj.bus_conductor

    def dehydrate_conductor_phone(self, obj):
        busfees_master_instance = busfees_master.objects.filter(bus_id=obj.bus_id).first()
        bus_master_instance = bus_master.objects.filter(bus_route=busfees_master_instance.route).first() if busfees_master_instance else None
        return bus_master_instance.conductor_phone if bus_master_instance else None
        # return obj.conductor_phone

    def dehydrate_bus_attendant(self, obj):
        busfees_master_instance = busfees_master.objects.filter(bus_id=obj.bus_id).first()
        bus_master_instance = bus_master.objects.filter(bus_route=busfees_master_instance.route).first() if busfees_master_instance else None
        return bus_master_instance.bus_attendant if bus_master_instance else None
        # return obj.bus_attendant

    def dehydrate_attendant_phone(self, obj):
        busfees_master_instance = busfees_master.objects.filter(bus_id=obj.bus_id).first()
        bus_master_instance = bus_master.objects.filter(bus_route=busfees_master_instance.route).first() if busfees_master_instance else None
        return bus_master_instance.attendant_phone if bus_master_instance else None


class TransportAdmin(ExportMixin,admin.ModelAdmin):
    resource_class = TransportMasterResource
    list_display = (
        'student_name', 'addmission_no', 'class_no', 'section', 'destination', 'route', 
        'father_name', 'phone_no', 'bus_driver', 'driver_phone', 'bus_conductor', 
        'conductor_phone', 'bus_attendant', 'attendant_phone'
    )

   

    list_filter = (BusRouteFilter, DestinationFilter, ClassFilter)

    def has_add_permission(self, request):
        return False

    def has_change_permission(self, request, obj=None):
        return False

    def has_delete_permission(self, request, obj=None):
        return False

        # Override get_search_results to handle custom filtering
    def get_search_results(self, request, queryset, search_term):
        bus_route = request.GET.get('bus_route', None)
        destination = request.GET.get('destination', None)
        class_no = request.GET.get('class_no', None)
        print('bus_route',bus_route)

         # If no filters are selected, return an empty queryset
        if not bus_route and not destination and not class_no:
            return queryset.none(), False

        # Annotate the queryset with related fields
        # latest_student_class = student_class.objects.filter(
        #     student_id=OuterRef('student_id')
        # ).order_by('-student_class_id')

        # bus_route_subquery = busfees_master.objects.filter(
        #     bus_id=OuterRef('bus_id')
        # ).values('route')[:1]

        # bus_destination_subquery = busfees_master.objects.filter(
        #     bus_id=OuterRef('bus_id')
        # ).values('destination')[:1]

        
        # bus_driver_subquery = bus_master.objects.filter(
        #     bus_route=OuterRef('route')
        # ).values('bus_driver')[:1]

        
        # driver_phone_subquery = bus_master.objects.filter(
        #     bus_route=OuterRef('route')
        # ).values('driver_phone')[:1]

        # bus_conductor_subquery = bus_master.objects.filter(
        #     bus_route=OuterRef('route')
        # ).values('bus_conductor')[:1]

        # conductor_phone_subquery = bus_master.objects.filter(
        #     bus_route=OuterRef('route')
        # ).values('conductor_phone')[:1]

        # bus_attendant_subquery = bus_master.objects.filter(
        #     bus_route=OuterRef('route')
        # ).values('bus_attendant')[:1]

        # attendant_phone_subquery = bus_master.objects.filter(
        #     bus_route=OuterRef('route')
        # ).values('attendant_phone')[:1]

        # queryset = queryset.annotate(
        #     class_no=Subquery(latest_student_class.values('class_no')[:1]),
        #     section=Subquery(latest_student_class.values('section')[:1]),
        #     route=Subquery(bus_route_subquery),
        #     destination=Subquery(bus_destination_subquery),
        #     bus_driver=Subquery(bus_driver_subquery),
        #     driver_phone=Subquery(driver_phone_subquery),
        #     bus_conductor=Subquery(bus_conductor_subquery),
        #     conductor_phone=Subquery(conductor_phone_subquery),
        #     bus_attendant=Subquery(bus_attendant_subquery),
        #     attendant_phone=Subquery(attendant_phone_subquery)
        # )

        # Filtering logic based on query parameters
        if bus_route:
            bus_ids = busfees_master.objects.filter(route=bus_route).values_list('bus_id', flat=True)
            queryset = queryset.filter(bus_id__in=bus_ids)
        if destination:
            bus_ids1 = busfees_master.objects.filter(destination=destination).values_list('bus_id', flat=True)
            # queryset = queryset.filter(destination=destination)
            queryset = queryset.filter(bus_id__in=bus_ids1)
        if class_no:
            student_ids = student_class.objects.filter(class_no=class_no).values_list('student_id', flat=True)
            queryset = queryset.filter(student_id__in=student_ids)
            # queryset = queryset.filter(class_no=class_no)

        return queryset, False


    # Override changelist_view to customize filters
    def changelist_view(self, request, extra_context=None):
        extra_context = extra_context or {}

        # Extract filter parameters from the request
        bus_route = request.GET.get('bus_route', '')
        destination = request.GET.get('destination', '')
        class_no = request.GET.get('class_no', '')

        # Add filter data to context for rendering the filters
        extra_context['bus_route_filter'] = range(1, 21)
        extra_context['class_filter'] = CLASS_CHOICES
        extra_context['destination_filter'] = busfees_master.objects.values_list('destination', flat=True).distinct()

        # Attach the filtered queryset to the changelist view
        response = super().changelist_view(request, extra_context=extra_context)
        
        return response


    # Customize the changelist template to include custom search fields
    change_list_template = 'admin/transport_changelist.html'

    # Custom URLs for AJAX
    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path('ajax/load-destinations/', self.admin_site.admin_view(self.load_destinations), name='ajax_load_destinations'),
        ]
        return custom_urls + urls

    # AJAX method to dynamically load destinations based on bus route
    def load_destinations(self, request):
        route = request.GET.get('route')
        destinations = list(busfees_master.objects.filter(route=route).values('destination').distinct())
        return JsonResponse(destinations, safe=False)

    # Custom JS for dynamic loading of destinations
    class Media:
        js = ('app/js/transport_filters.js',)

    # Define the remaining list_display methods (e.g., class_no, section)
    # def class_no(self, obj):
    #     return obj.class_no

    # def section(self, obj):
    #     return obj.section
    
    def class_no(self, obj):
        student_class_instance = student_class.objects.filter(student_id=obj.student_id).order_by('-started_on').first()
        return student_class_instance.class_no if student_class_instance else None
    # get_class_no.short_description = 'Class'

    def section(self, obj):
        student_class_instance = student_class.objects.filter(student_id=obj.student_id).order_by('-started_on').first()
        return student_class_instance.section if student_class_instance else None
    # get_section.short_description = 'Section'

    

    # def route(self, obj):
    #     return obj.route

    # def destination(self, obj):
    #     return obj.destination

    def route(self, obj):
        busfees_master_instance = busfees_master.objects.filter(bus_id=obj.bus_id).first()
        return busfees_master_instance.route if busfees_master_instance else None
        # return obj.route

    def destination(self, obj):
        busfees_master_instance = busfees_master.objects.filter(bus_id=obj.bus_id).first()
        return busfees_master_instance.destination if busfees_master_instance else None
        # return obj.destination

    # def bus_driver(self, obj):
    #     return obj.bus_driver
    
    # def driver_phone(self, obj):
    #     return obj.driver_phone

    def bus_driver(self, obj):
        busfees_master_instance = busfees_master.objects.filter(bus_id=obj.bus_id).first()
        bus_master_instance = bus_master.objects.filter(bus_route=busfees_master_instance.route).first() if busfees_master_instance else None
        return bus_master_instance.bus_driver if bus_master_instance else None
        # return obj.bus_driver
    
    def driver_phone(self, obj):
        busfees_master_instance = busfees_master.objects.filter(bus_id=obj.bus_id).first()
        bus_master_instance = bus_master.objects.filter(bus_route=busfees_master_instance.route).first() if busfees_master_instance else None
        return bus_master_instance.driver_phone if bus_master_instance else None

        # return obj.driver_phone

    # def bus_conductor(self, obj):
    #     return obj.bus_conductor

    # def conductor_phone(self, obj):
    #     return obj.conductor_phone

    # def bus_attendant(self, obj):
    #     return obj.bus_attendant

    # def attendant_phone(self, obj):
    #     return obj.attendant_phone

    def bus_conductor(self, obj):
        busfees_master_instance = busfees_master.objects.filter(bus_id=obj.bus_id).first()
        bus_master_instance = bus_master.objects.filter(bus_route=busfees_master_instance.route).first() if busfees_master_instance else None
        return bus_master_instance.bus_conductor if bus_master_instance else None
        # return obj.bus_conductor

    def conductor_phone(self, obj):
        busfees_master_instance = busfees_master.objects.filter(bus_id=obj.bus_id).first()
        bus_master_instance = bus_master.objects.filter(bus_route=busfees_master_instance.route).first() if busfees_master_instance else None
        return bus_master_instance.conductor_phone if bus_master_instance else None
        # return obj.conductor_phone

    def bus_attendant(self, obj):
        busfees_master_instance = busfees_master.objects.filter(bus_id=obj.bus_id).first()
        bus_master_instance = bus_master.objects.filter(bus_route=busfees_master_instance.route).first() if busfees_master_instance else None
        return bus_master_instance.bus_attendant if bus_master_instance else None
        # return obj.bus_attendant

    def attendant_phone(self, obj):
        busfees_master_instance = busfees_master.objects.filter(bus_id=obj.bus_id).first()
        bus_master_instance = bus_master.objects.filter(bus_route=busfees_master_instance.route).first() if busfees_master_instance else None
        return bus_master_instance.attendant_phone if bus_master_instance else None
        # return obj.attendant_phone

    class_no.short_description = 'Class No'
    section.short_description = 'Section'
    route.short_description = 'Route'
    destination.short_description = 'Destination'
    bus_driver.short_description = 'Bus Driver'
    driver_phone.short_description = 'Driver Phone'
    bus_conductor.short_description = 'Bus Conductor'
    conductor_phone.short_description = 'Conductor Phone'
    bus_attendant.short_description = 'Bus Attendant'
    attendant_phone.short_description = 'Attendant Phone'

admin.site.register(transport, TransportAdmin)

class PassedOutFilter(admin.SimpleListFilter):
    title = 'PassedOut'
    parameter_name = 'radioval'

    def lookups(self, request, model_admin):
        return []

class YearFilter(admin.SimpleListFilter):
    title = 'Year'
    parameter_name = 'year'

    def lookups(self, request, model_admin):
        return []
    

class TuitionFeesDefaulterResource(resources.ModelResource):

    class Meta:
        model = tuition_fees_defaulter
        fields = ('student_name', 'admission_no', 'class_no', 'section', 'tmpval')

    student_name = Field(attribute='student_name', column_name='Student Name')
    addmission_no = Field(attribute='addmission_no', column_name='Admission No')
    class_no = Field(attribute='class_no', column_name='Class')
    section = Field(attribute='section', column_name='Section')
    tmpval = Field(attribute='tmpval', column_name='Tution Fees unpaid for months')

    def dehydrate_class_no(self, obj):
        # Assuming 'bus_driver' is a related field on the model
        student_class_instance = student_class.objects.filter(student_id=obj.student_id.student_id).order_by('-started_on').first()
        return student_class_instance.class_no if student_class_instance else None

    
    def dehydrate_section(self, obj):
        student_class_instance = student_class.objects.filter(student_id=obj.student_id.student_id).order_by('-started_on').first()
        return student_class_instance.section if student_class_instance else None

    def dehydrate_tmpval(self, obj):
        # Get the year from the object
        year = int(obj.year)
        month_array = get_months_array(year)

        # Calculate the unpaid months
        unpaid_months = calculate_unpaid_months(obj.months_paid, month_array)

        # If unpaid_months is an empty string, return None to exclude it from display
        if not unpaid_months:
            return None  # or you can return '' if you prefer an empty string to be displayed
        
        return unpaid_months


    



class TuitionFeesDefaulterAdmin(ExportMixin,admin.ModelAdmin):
    resource_class = TuitionFeesDefaulterResource
    list_display = ('admission_no', 'student_name', 'class_no', 'section', 'tmpval')
    list_filter = (PassedOutFilter, ClassFilter, YearFilter)

    def has_add_permission(self, request):
        return False

    def has_change_permission(self, request, obj=None):
        return False

    def has_delete_permission(self, request, obj=None):
        return False
    
    def get_year(self, request):
        current_month = date.today().month
        return date.today().year if current_month >= 4 else date.today().year - 1


    def changelist_view(self, request, extra_context=None):
        year_choices = [str(year) for year in range(2024, 2017, -1)]
        extra_context = extra_context or {}
        extra_context['class_choices'] = CLASS_CHOICES
        extra_context['year_choices'] = year_choices
        return super().changelist_view(request, extra_context=extra_context)

    change_list_template = 'admin/tuitionfeesdefaulter_change_list.html'

    def get_search_results(self, request, queryset, search_term):
        radioval = request.GET.get('radioval', None)
        year1 = request.GET.get('year', None)
        class_no = request.GET.get('class_no', None)

        year = int(year1) if year1 else self.get_year(request)
        month_array = get_months_array(year)
        months_str = ','.join(map(str, month_array))

        if not radioval and not year1 and not class_no:
            return queryset.none(), False

        total_fees_paid = ExpressionWrapper(
            Coalesce(F('tuition_fees_paid'), 0) +
            Coalesce(F('funds_fees_paid'), 0) +
            Coalesce(F('sports_fees_paid'), 0),
            output_field=DecimalField()
        )

        queryset = queryset.annotate(
            total_fees_paid=total_fees_paid
        ).filter(
            year=year,
            total_fees_paid__gt=0
        ).annotate(
            months_paid=RawSQL(
                """
                SELECT GROUP_CONCAT(DISTINCT TRIM(sf.fees_period_month) 
                ORDER BY sf.fees_period_month+0 SEPARATOR ', ')
                FROM student_fees AS sf
                WHERE sf.student_id = student_fees.student_id
                AND sf.year = %s
                """, (year,)
            ),
            student_name=F('student_id__student_name'),
            admission_no=F('student_id__addmission_no'),
            class_no=F('student_class'),
            section=F('student_section'),
            passedout_date=F('student_id__passedout_date'),
            tmpval=Value('', output_field=CharField())  # Placeholder for tmpval logic
        ).exclude(
            months_paid=months_str
        )

        if class_no:
            queryset = queryset.filter(class_no=class_no)

        queryset = self.filter_passed_out_students(queryset, radioval)
        for obj in queryset:
            obj.tmpval = calculate_unpaid_months(obj.months_paid, month_array)

        # Filter out objects where `tmpval` is empty
        filtered_queryset = [obj.student_fee_id for obj in queryset if obj.tmpval == '']

        queryset = queryset.exclude(student_fee_id__in=filtered_queryset)

        return queryset, False

    # def get_months_array(self, year):
    #     start_date = date(year, 4, 1)
    #     end_date = date(year + 1, 3, 31)
    #     current_date = date.today()

    #     months = []
    #     time = start_date
    #     if current_date < end_date:
    #         end_date = current_date

    #     while time <= end_date:
    #         months.append(time.month)
    #         next_month = (time.month % 12) + 1
    #         next_year = time.year + (1 if next_month == 1 else 0)
    #         time = date(next_year, next_month, 1)

    #     return months

    # def calculate_unpaid_months(self, paid_months, months_array):
    #     print('paid_months in calculate_unpaid_months', paid_months)
    #     print('months_array in calculate_unpaid_months', months_array)
    #     if not paid_months:
    #         return ','.join(map(str, months_array))  # Return all months if no months are paid

    #     paid_months_list = list(map(int, paid_months.split(',')))
    #     unpaid_months = set(months_array) - set(paid_months_list)
    #     return ','.join(map(str, sorted(unpaid_months)))

    def filter_passed_out_students(self, queryset, radioval):
        current_date = date.today()
        if radioval != 'withpassedout':
            queryset = queryset.exclude(passedout_date__lt=current_date)
        return queryset
    
    
#     # Display functions for fields
    def student_name(self, obj):
        return obj.student_name

    def admission_no(self, obj):
        return obj.admission_no

    def class_no(self, obj):
        return obj.class_no

    def section(self, obj):
        return obj.section

    def tmpval(self, obj):
        # Get the year from the object
        year = int(obj.year)
        month_array = get_months_array(year)

        # Calculate the unpaid months
        unpaid_months = calculate_unpaid_months(obj.months_paid, month_array)

        # If unpaid_months is an empty string, return None to exclude it from display
        if not unpaid_months:
            return None  # or you can return '' if you prefer an empty string to be displayed
        
        obj.tmpval = unpaid_months 
        return unpaid_months
    tmpval.short_description = 'Tuition Fees Unpaid for Months'


admin.site.register(tuition_fees_defaulter, TuitionFeesDefaulterAdmin)

# Custom Filter for Class No
class ClassNoFilter1(admin.SimpleListFilter):
    title = 'Class'
    parameter_name = 'class_no'

    def lookups(self, request, model_admin):
        # Define class choices as per your requirements
        return CLASS_CHOICES

    def queryset(self, request, queryset):
        cls = self.value()
        if cls:
            subquery = student_class.objects.filter(
                student_id=OuterRef('student_id'),
                class_no=cls
            ).order_by('-student_class_id').values('student_class_id')[:1]
            return queryset.filter(
                student_id__in=Subquery(subquery)
            )
        return queryset


    
# Custom Filter for Admission Date From
class DateFromFilter(admin.SimpleListFilter):
    title = 'Admission Date From'
    parameter_name = 'date_from'

    def lookups(self, request, model_admin):
        # return ()
        return []

    # def queryset(self, request, queryset):
    #     date_from = request.GET.get('date_from')
    #     if date_from:
    #         return queryset.filter(admission_date__gte=date_from)
    #     return queryset

    # def choices(self, changelist):
    #     # Override to remove the 'All' option
    #     return []

# Custom Filter for Admission Date To
class DateToFilter(admin.SimpleListFilter):
    title = 'Admission Date To'
    parameter_name = 'date_to'

    def lookups(self, request, model_admin):
        # return ()
        return []

    # def queryset(self, request, queryset):
    #     date_to = request.GET.get('date_to')
    #     if date_to:
    #         return queryset.filter(admission_date__lte=date_to)
    #     return queryset

    # def choices(self, changelist):
    #     # Override to remove the 'All' option
    #   
    # 


class AdmissionReportResource(resources.ModelResource):
    # def __init__(self, request=None, *args, **kwargs):
    #     super().__init__(*args, **kwargs)
    #     self._request = request  # Store the request passed from the Admin class
    #     # Handle other keyword arguments like 'encoding'
    #     # self.extra_args = kwargs

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self._request = None  # Default to None; it will be set later if passed
    

    class Meta:
        model = admission_report
        fields = ('addmission_no', 'admission_date', 'student_name', 'birth_date',
                  'class_no', 'section', 'father_name', 'mother_name', 'address')

    # Define fields
    student_name = Field(attribute='student_name', column_name='Student Name')
    addmission_no = Field(attribute='addmission_no', column_name='Admission No')
    admission_date = Field(attribute='admission_date', column_name='Admission Date')
    class_no = Field(attribute='class_no', column_name='Class')
    section = Field(attribute='section', column_name='Section')
    father_name = Field(attribute='father_name', column_name='Father Name')
    mother_name = Field(attribute='phone_no', column_name='Mother Name')
    address = Field(attribute='address', column_name='Address')
    birth_date = Field(attribute='birth_date', column_name='Birth Date')

    # Custom dehydrate methods
    # def dehydrate_class_no(self, obj):
    #     search_class_no = self._request.GET.get('class_no', None) if self._request else None
    #     print('search_class_no',search_class_no)
    #     if search_class_no:
    #         student_class_instance = student_class.objects.filter(
    #             student_id=obj.student_id, class_no=search_class_no).order_by('-started_on').first()
    #     else:
    #         student_class_instance = student_class.objects.filter(
    #             student_id=obj.student_id).order_by('-started_on').first()

    #     return student_class_instance.class_no if student_class_instance else None

    # def dehydrate_section(self, obj):
    #     search_class_no = self._request.GET.get('class_no', None) if self._request else None

    #     if search_class_no:
    #         student_class_instance = student_class.objects.filter(
    #             student_id=obj.student_id, class_no=search_class_no).order_by('-started_on').first()
    #     else:
    #         student_class_instance = student_class.objects.filter(
    #             student_id=obj.student_id).order_by('-started_on').first()

    #     return student_class_instance.section if student_class_instance else None


    def dehydrate_class_no(self, obj):
        # Check if request is available
        if self._request:
            search_class_no = self._request.GET.get('class_no', None)
        else:
            search_class_no = None

        # Filter by class_no if available
        if search_class_no:
            student_class_instance = student_class.objects.filter(
                student_id=obj.student_id, class_no=search_class_no).order_by('-started_on').first()
        else:
            student_class_instance = student_class.objects.filter(
                student_id=obj.student_id).order_by('-started_on').first()

        return student_class_instance.class_no if student_class_instance else None

    def dehydrate_section(self, obj):
        # Check if request is available
        if self._request:
            search_class_no = self._request.GET.get('class_no', None)
        else:
            search_class_no = None

        # Filter by class_no and return section if available
        if search_class_no:
            student_class_instance = student_class.objects.filter(
                student_id=obj.student_id, class_no=search_class_no).order_by('-started_on').first()
        else:
            student_class_instance = student_class.objects.filter(
                student_id=obj.student_id).order_by('-started_on').first()

        return student_class_instance.section if student_class_instance else None

class AdmissionReportAdmin(ExportMixin, admin.ModelAdmin):

    resource_class = AdmissionReportResource
    list_display = (
        'addmission_no', 'formatted_admission_date', 'student_name', 'birth_date',
        'class_no', 'section', 'father_name', 'mother_name', 'address'
    )
    # search_fields = ('student_name', 'admission_no')
    list_filter = (ClassFilter, DateFromFilter, DateToFilter)

    def has_add_permission(self, request):
        return False

    def has_change_permission(self, request, obj=None):
        return False

    def has_delete_permission(self, request, obj=None):
        return False
    
    

    def changelist_view(self, request, extra_context=None):
        """
        Handle the custom filters and pass them to the context for rendering the change list view.
        """
        extra_context = extra_context or {}
        # Store the request object in the instance
        
        # Get the filter values from GET parameters
        class_no = request.GET.get('class_no', '')
        date_from = request.GET.get('date_from', '')
        date_to = request.GET.get('date_to', timezone.now().strftime('%Y-%m-%d'))

        # Pass filter values to the template to retain them
        extra_context['class_no'] = class_no
        extra_context['date_from'] = date_from
        extra_context['date_to'] = date_to

        # Retrieve distinct class numbers from student_class to populate the filter options
        distinct_classes = student_class.objects.values_list('class_no', flat=True).distinct()
        # extra_context['class_choices'] = [(cls, cls) for cls in distinct_classes]
        extra_context['class_choices'] = CLASS_CHOICES
        self._request = request
        # Call the default changelist_view with the extra context
        return super().changelist_view(request, extra_context=extra_context)
    
    change_list_template = 'admin/admissionreport_change_list.html'

    # def get_export_resource_class(self):
    #     """
    #     Pass the stored request object to the ModelResource class constructor.
    #     """
    #     # Return an instance of the resource class with request passed
    #     return self.resource_class

    # def get_export_resource_instance(self):
    #     """
    #     Pass the stored request object when instantiating the resource class.
    #     """
    #     # Debugging: Log the request being passed
    #     print("Passing request to resource:", self._request.GET)  
    #     return self.resource_class(request=self._request)  
    def get_export_resource_instance(self):
        """
        Pass the stored request object when instantiating the resource class.
        """
        resource = self.resource_class()  # Create an instance of the resource class
        # Inject the request into the resource class
        if hasattr(self, '_request'):
            resource._request = self._request  # Set the _request attribute
        return resource
    # def get_search_results(self, request, queryset, search_term):
    #     """
    #     Override get_search_results to filter by class_no and date range.
    #     """
    #     queryset, use_distinct = super().get_search_results(request, queryset, search_term)

    #     # Get custom filter parameters from the request
    #     class_no = request.GET.get('class_no', '')
    #     date_from = request.GET.get('date_from', '')
    #     date_to = request.GET.get('date_to', timezone.now().strftime('%Y-%m-%d'))

    #     # Apply class_no filter if provided
    #     if class_no:
    #         subquery = student_class.objects.filter(
    #             student_id=OuterRef('student_id'),
    #             class_no=class_no
    #         ).order_by('-student_class_id').values('student_class_id')[:1]
    #         queryset = queryset.filter(student_id__in=Subquery(subquery))

    #     # Apply date filters if provided
    #     if date_from:
    #         queryset = queryset.filter(admission_date__gte=date_from)
    #     if date_to:
    #         queryset = queryset.filter(admission_date__lte=date_to)

    #     return queryset, use_distinct


    # def get_search_results(self, request, queryset, search_term):
    #     """
    #     Override get_search_results to filter by class_no and date range.
    #     """
    #     queryset, use_distinct = super().get_search_results(request, queryset, search_term)

    #     # Get custom filter parameters from the request
    #     class_no = request.GET.get('class_no', '')
    #     date_from = request.GET.get('date_from', '')
    #     date_to = request.GET.get('date_to', timezone.now().strftime('%Y-%m-%d'))

    #     # Apply class_no filter if provided using Exists() instead of Subquery
    #     if class_no:
    #         student_class_qs = student_class.objects.filter(
    #             student_id=OuterRef('student_id'),
    #             class_no=class_no
    #         ).order_by('-student_class_id')

    #         queryset = queryset.annotate(class_exists=Exists(student_class_qs)).filter(class_exists=True)

    #     # Apply date filters if provided
    #     if date_from:
    #         queryset = queryset.filter(admission_date__gte=date_from)
    #     if date_to:
    #         queryset = queryset.filter(admission_date__lte=date_to)

    #     return queryset, use_distinct


    def get_search_results(self, request, queryset, search_term):
        """
        Override get_search_results to filter by class_no and date range.
        """
        queryset, use_distinct = super().get_search_results(request, queryset, search_term)

        # Get custom filter parameters from the request
        class_no = request.GET.get('class_no', '')
        date_from = request.GET.get('date_from', '')
        date_to = request.GET.get('date_to', timezone.now().strftime('%Y-%m-%d'))

        if not class_no and not date_from:
            return queryset.none(), use_distinct

        # Apply strict class_no filter if provided using Exists() with exact class_no match
        if class_no:
            student_class_qs = student_class.objects.filter(
                student_id=OuterRef('student_id'),
                class_no=class_no
            ).order_by('-student_class_id')  # Ensure you're checking the latest class

            queryset = queryset.annotate(class_exists=Exists(student_class_qs)).filter(class_exists=True)

        # Apply date filters if provided
        if date_from:
            queryset = queryset.filter(admission_date__gte=date_from)
        if date_to:
            queryset = queryset.filter(admission_date__lte=date_to)

        return queryset, use_distinct

    def formatted_admission_date(self, obj):
        # Format the admission date in dd-mm-yyyy
        return obj.admission_date.strftime('%d-%m-%Y')
    formatted_admission_date.short_description = 'Admission Date'

    def class_no(self, obj):
        # Subquery to fetch class_no from student_class without foreign key
        # subquery = student_class.objects.filter(
        #     student_id=obj.student_id
        # ).order_by('-student_class_id').values('class_no')[:1]
        # result = subquery.first()
        # return result.get('class_no') if result else None
        search_class_no = self._request.GET.get('class_no', None)

        # Filter by class_no and/or section if available
        if search_class_no:
            student_class_instance = student_class.objects.filter(student_id=obj.student_id, class_no=search_class_no).order_by('-started_on').first()
        else:
            student_class_instance = student_class.objects.filter(student_id=obj.student_id).order_by('-started_on').first()

        return student_class_instance.class_no if student_class_instance else None

    def section(self, obj):
        # Subquery to fetch section from student_class without foreign key
        # subquery = student_class.objects.filter(
        #     student_id=obj.student_id
        # ).order_by('-student_class_id').values('section')[:1]
        # result = subquery.first()
        # return result.get('section') if result else None
        search_class_no = self._request.GET.get('class_no', None)

        # Filter by class_no and/or section if available
        if search_class_no:
            student_class_instance = student_class.objects.filter(student_id=obj.student_id, class_no=search_class_no).order_by('-started_on').first()
        else:
            student_class_instance = student_class.objects.filter(student_id=obj.student_id).order_by('-started_on').first()

        return student_class_instance.section if student_class_instance else None

# Register the AdmissionReport proxy model with the custom admin
admin.site.register(admission_report, AdmissionReportAdmin)


from django.shortcuts import render
from django.template.response import TemplateResponse
from django.utils.datetime_safe import datetime


class DecimalEncoder(json.JSONEncoder):
    def default(self, obj):
        if isinstance(obj, Decimal):
            return str(obj)  # Convert to string
        return super().default(obj)
    
class FinalFeesReportAdmin(admin.ModelAdmin):
    model = student_fee
    change_list_template = "admin/final_fees_report_change_list.html"  # Custom template
    
    # list_display = ('student_class',)
    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path('export-final-fees-report/', self.admin_site.admin_view(self.export_final_fees_report_to_excel), name='export_final_fees_report'),
        ]
        return custom_urls + urls
    
    # Permissions removed
    def has_add_permission(self, request):
        return False

    def has_change_permission(self, request, obj=None):
        return False

    def has_delete_permission(self, request, obj=None):
        return False  
    

    def changelist_view(self, request, extra_context=None):
        cheque_details = []
        reporttype = ''
        year = ''
        total_annual_fees = 0
        total_tuition_fees = 0
        total_funds_fees = 0
        total_sports_fees = 0
        total_activity_fees = 0
        total_admission_fees = 0
        total_security_fees = 0
        total_late_fees = 0
        total_dayboarding_fees = 0
        total_bus_fees = 0
        total_fees_amount = 0
        total_amount_paid = 0
        if request.method == 'POST':
            # reporttype = request.GET.get('reporttype', 'summaryreport')  # Default to 'summaryreport'
            # year = request.GET.get('year', datetime.now().year)  # Default to current year

            reporttype = request.POST.get('reporttype')  # Default to 'summaryreport'
            year = request.POST.get('year') 
            print('reporttype',reporttype)
            
            extra_context = extra_context or {}
            extra_context['reporttype'] = reporttype
            extra_context['year'] = year
            extra_context['years'] = list(range(2000, 2026))[::-1]  # Populate years 2000-2025
            extra_context['report_types'] = ["Summary report","Actual report", "Advance report"]  # Populate report types
            # extra_context['opts'] = self.model._meta
            extra_context['cheque_details'] = cheque_details
            extra_context['total_annual_fees'] = total_annual_fees
            extra_context['total_tuition_fees'] = total_tuition_fees
            extra_context['total_funds_fees'] = total_funds_fees
            extra_context['total_sports_fees'] = total_sports_fees
            extra_context['total_activity_fees'] = total_activity_fees


            extra_context['total_admission_fees'] = total_admission_fees
            extra_context['total_security_fees'] = total_security_fees
            extra_context['total_late_fees'] = total_late_fees
            extra_context['total_dayboarding_fees'] = total_dayboarding_fees
            extra_context['total_bus_fees'] = total_bus_fees

            extra_context['total_fees_amount'] = total_fees_amount
            extra_context['total_amount_paid'] = total_amount_paid
            extra_context['export_url'] = 'export-final-fees-report/'
            if (reporttype == "") and not year:
                messages.error(request, _("Both 'reporttype' and 'year' fields are required."))
                return super().changelist_view(request, extra_context=extra_context) 
            # Ensure year is valid
            try:
                year = int(year)
            except ValueError:
                # year = datetime.now().year
                messages.error(request, _("Invalid year value provided."))
                return super().changelist_view(request, extra_context=extra_context)

           

            if reporttype == "summary-report":
                to_year = year + 1
                from_date = f"{year}-04-01"
                to_date = f"{to_year}-03-31"

                # Write the raw SQL query
                sql_query = f"""
                    SELECT student_class,
                        SUM(annual_fees_paid) AS sum_annual_fees_paid,
                        SUM(tuition_fees_paid) AS sum_tuition_fees_paid,
                        SUM(funds_fees_paid) AS sum_funds_fees_paid,
                        SUM(sports_fees_paid) AS sum_sports_fees_paid,
                        SUM(activity_fees) AS sum_activity_fees,
                        SUM(admission_fees_paid) AS sum_admission_fees_paid,
                        SUM(security_paid) AS sum_security_paid,
                        SUM(late_fees_paid) AS sum_late_fees_paid,
                        SUM(dayboarding_fees_paid) AS sum_dayboarding_fees_paid,
                        SUM(bus_fees_paid) AS sum_bus_fees_paid,
                        SUM(Total_amount) AS sum_total_amount,
                        SUM(amount_paid) AS sum_amount_paid
                    FROM student_fees
                    WHERE date_payment BETWEEN '{from_date}' AND '{to_date}'
                    GROUP BY student_class
                    ORDER BY student_class + 0 ASC
                """

                # Execute the raw SQL query
                with connection.cursor() as cursor:
                    cursor.execute(sql_query)
                    rows = cursor.fetchall()

                print(f'Query Results: {rows}')

                # Build the string output similar to the PHP loop

                for row in rows:
                    cheque_details.append({
                        'student_class': row[0],
                        'total_annual_fees': row[1],
                        'total_tuition_fees': row[2],
                        'total_funds_fees': row[3],
                        'total_sports_fees': row[4],  # Corrected the mobile number index
                        'total_activity_fees': row[5],
                        'total_admission_fees': row[6],
                        'total_security_fees': row[7],
                        'total_late_fees': row[8],
                        'total_dayboarding_fees': row[9],
                        'total_bus_fees': row[10],
                        'total_fees_amount': row[11],
                        'total_amount_paid': row[12],
                    })
                    
                    total_annual_fees += row[1] if row[1] else 0
                    total_tuition_fees += row[2] if row[2] else 0
                    total_funds_fees += row[3] if row[3] else 0
                    total_sports_fees += row[4] if row[4] else 0
                    total_activity_fees += row[5] if row[5] else 0
                    total_admission_fees += row[6] if row[6] else 0
                    total_security_fees += row[7] if row[7] else 0
                    total_late_fees += row[8] if row[8] else 0
                    total_dayboarding_fees += row[9] if row[9] else 0
                    total_bus_fees += row[10] if row[10] else 0
                    total_fees_amount += row[11] if row[11] else 0
                    total_amount_paid += row[12] if row[12] else 0
                    # sumdetails += ",".join([str(item) for item in row]) + "$"


        # queryset = self.get_filtered_queryset(reporttype, year)

        extra_context = extra_context or {}
        extra_context['reporttype'] = reporttype
        extra_context['year'] = year
        extra_context['years'] = list(range(2000, 2026))[::-1]  # Populate years 2000-2025
        extra_context['report_types'] = ["Summary report","Actual report", "Advance report"]  # Populate report types
        # extra_context['opts'] = self.model._meta
        request.session['cheque_details'] = json.dumps(cheque_details, cls=DecimalEncoder)
        extra_context['cheque_details'] = cheque_details
        extra_context['total_annual_fees'] = total_annual_fees
        extra_context['total_tuition_fees'] = total_tuition_fees
        extra_context['total_funds_fees'] = total_funds_fees
        extra_context['total_sports_fees'] = total_sports_fees
        extra_context['total_activity_fees'] = total_activity_fees


        extra_context['total_admission_fees'] = total_admission_fees
        extra_context['total_security_fees'] = total_security_fees
        extra_context['total_late_fees'] = total_late_fees
        extra_context['total_dayboarding_fees'] = total_dayboarding_fees
        extra_context['total_bus_fees'] = total_bus_fees

        extra_context['total_fees_amount'] = total_fees_amount
        extra_context['total_amount_paid'] = total_amount_paid
        extra_context['export_url'] = 'export-final-fees-report/'

        return super().changelist_view(request, extra_context=extra_context)

    def export_final_fees_report_to_excel(self, request):
        # Assuming `cheque_details` is the list of objects representing the table data
        # cheque_details = request.session.get('cheque_details', [])
        cheque_details = json.loads(request.session['cheque_details'])

        if not cheque_details:
            # Handle the case where there is no data to export
            messages.error(request, "No data available to export.")
            return redirect('admin:final_fees_report_changelist')

        # Define the column headers as per the table in the image
        # headers = [
        #     'Student Class', 'Annual Fees', 'Tuition Fees', 'Funds Fees', 'Sports Fees', 
        #     'Activity Fees', 'Admission Fees', 'Security Fees', 'Late Fees', 
        #     'Miscellaneous Fees', 'Bus Fees', 'Total Fees', 'Amount Paid after Concession'
        # ]
        headers = [
            'Class', 'Annual', 'Tuition', 'Funds', 'Sports', 
            'Activity', 'Admission', 'Security', 'Late', 
            'Dayboarding', 'Bus', 'Total', 'Total fees after concession'
        ]

        # Convert the list of objects to a DataFrame for export
        data = []
        for obj in cheque_details:
            data.append([
                obj['student_class'], Decimal(obj['total_annual_fees']) if obj['total_annual_fees'] else 0, Decimal(obj['total_tuition_fees']) if obj['total_tuition_fees'] else 0,
                Decimal(obj['total_funds_fees']) if obj['total_funds_fees'] else 0, Decimal(obj['total_sports_fees']) if obj['total_sports_fees'] else 0, Decimal(obj['total_activity_fees']) if obj['total_activity_fees'] else 0,
                Decimal(obj['total_admission_fees']) if obj['total_admission_fees'] else 0, Decimal(obj['total_security_fees']) if obj['total_security_fees'] else 0, Decimal(obj['total_late_fees']) if obj['total_late_fees'] else 0,
                Decimal(obj['total_dayboarding_fees']) if obj['total_dayboarding_fees'] else 0, Decimal(obj['total_bus_fees']) if obj['total_bus_fees'] else 0, Decimal(obj['total_fees_amount']),
                Decimal(obj['total_amount_paid'])
            ])

        # Create DataFrame from data
        df = pd.DataFrame(data, columns=headers)

        # Add a total row at the end if necessary
        total_row = [
            'Total', sum(df['Annual Fees']), sum(df['Tuition Fees']), sum(df['Funds Fees']),
            sum(df['Sports Fees']), sum(df['Activity Fees']), sum(df['Admission Fees']),
            sum(df['Security Fees']), sum(df['Late Fees']), sum(df['Miscellaneous Fees']),
            sum(df['Bus Fees']), sum(df['Total Fees']), sum(df['Amount Paid after Concession'])
        ]
        df.loc[len(df.index)] = total_row

        # Create a BytesIO stream to hold the Excel data
        excel_file = BytesIO()
        with pd.ExcelWriter(excel_file, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Final Fees Report')

        # Rewind the buffer
        excel_file.seek(0)

        # Create the HTTP response to download the file
        response = HttpResponse(
            excel_file.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=final_fees_report.xlsx'

        return response

admin.site.register(final_fees_report, FinalFeesReportAdmin)




class TransportDefaulterResource(resources.ModelResource):

    class Meta:
        model = transport_defaulter
        fields = ('addmission_no','student_name', 'student_class', 'student_section', 
        'route', 'destination', 'unpaid_months')

    addmission_no = Field(attribute='addmission_no', column_name='Admission No')
    student_name = Field(attribute='student_name', column_name='Student Name')
    student_class = Field(attribute='student_class', column_name='Class')
    student_section = Field(attribute='student_section', column_name='Section')
    route = Field(attribute='route', column_name='Route')
    destination = Field(attribute='destination', column_name='Destination')
    unpaid_months = Field(attribute='unpaid_months', column_name='Busfees unpaid for months')

    def dehydrate_student_name(self, obj):
        return obj.student_id.student_name

    
    def dehydrate_addmission_no(self, obj):
        return obj.student_id.addmission_no
    
    def dehydrate_unpaid_months(self, obj):
        # Get the year from the object
        year = int(obj.year)
        current_date = datetime.now().date()
        # year = self.get_year()
        montharray = get_months_array(year)

        months_paid = obj.months_paid
        passedout_date = obj.student_id.passedout_date

        if not months_paid:
            months_paid_list = []
        else:
            try:
                # Remove duplicates and convert months_paid_list to integers
                months_paid_list = list(map(int, set(map(str.strip, months_paid.split(',')))))
            except ValueError as e:
                print(f"Error converting months_paid to integers: {e}")
                months_paid_list = []

        # Calculate unpaid months
        unpaid_months = list(set(montharray) - set(months_paid_list))

        # Only show unpaid months if student hasn't passed out
        if unpaid_months and (not passedout_date or passedout_date >= current_date):
            return ",".join(map(str, unpaid_months))
        else:
            return "No unpaid months"



class TransportDefaulterReportAdmin(ExportMixin, admin.ModelAdmin):
    resource_class = TransportDefaulterResource
    # Define the fields to display in the admin list
    list_display = (
        'admission_no','student_name', 'class_no', 'section', 
        'route', 'destination', 'unpaid_months'
    )
    
    list_filter = (ClassFilter, YearFilter, BusRouteFilter, DestinationFilter)
    
    # Disable add, change, and delete permissions
    def has_add_permission(self, request):
        return False

    def has_change_permission(self, request, obj=None):
        return False

    def has_delete_permission(self, request, obj=None):
        return False

    # Custom method to show the student's name
    def student_name(self, obj):
        return obj.student_id.student_name

    # Custom method to show the student's admission number
    def admission_no(self, obj):
        return obj.student_id.addmission_no

    # Custom method to show the student's class
    def class_no(self, obj):
        return obj.student_class

    # Custom method to show the student's section
    def section(self, obj):
        return obj.student_section

    # Custom method to show the route
    def route(self, obj):
        return obj.route

    # Custom method to show the destination
    def destination(self, obj):
        return obj.destination

    # # Custom method to show unpaid months
    def unpaid_months(self, obj):
        # Calculate the unpaid months
        current_date = datetime.now().date()
        year = self.get_year()
        montharray = get_months_array(year)

        months_paid = obj.months_paid
        passedout_date = obj.student_id.passedout_date

        if not months_paid:
            months_paid_list = []
        else:
            try:
                # Remove duplicates and convert months_paid_list to integers
                months_paid_list = list(map(int, set(map(str.strip, months_paid.split(',')))))
            except ValueError as e:
                print(f"Error converting months_paid to integers: {e}")
                months_paid_list = []

        # Calculate unpaid months
        unpaid_months = list(set(montharray) - set(months_paid_list))

        # Only show unpaid months if student hasn't passed out
        if unpaid_months and (not passedout_date or passedout_date >= current_date):
            return ",".join(map(str, unpaid_months))
        else:
            return "No unpaid months"

    
    # def changelist_view(self, request, extra_context=None):
    #     bus_route = request.GET.get('bus_route', '')
    #     destination = request.GET.get('destination', '')
    #     class_no = request.GET.get('class_no', '')

    #     year_choices = [str(year) for year in range(2024, 2017, -1)]
    #     extra_context = extra_context or {}
    #     extra_context['class_choices'] = CLASS_CHOICES
    #     extra_context['year_choices'] = year_choices
    #     extra_context['bus_route_choices'] = range(1, 21)
    #     extra_context['destination_choices'] = busfees_master.objects.values_list('destination', flat=True).distinct()
    #     extra_context['selected_bus_route'] = bus_route
    #     extra_context['selected_destination'] = destination
    #     extra_context['selected_class_no'] = class_no
      

    #     return super().changelist_view(request, extra_context=extra_context)
    # Override changelist_view to provide extra context for filters
    def changelist_view(self, request, extra_context=None):
        bus_route = request.GET.get('bus_route', '')
        destination = request.GET.get('destination', '')
        class_no = request.GET.get('class_no', '')
        year = request.GET.get('year', None)
        if year is None:
            year = self.get_year()

        year_choices = [str(year) for year in range(2024, 2017, -1)]
        extra_context = extra_context or {}
        extra_context['class_choices'] = CLASS_CHOICES
        extra_context['year_choices'] = year_choices
        extra_context['bus_route_choices'] = range(1, 21)
        extra_context['destination_choices'] = busfees_master.objects.values_list('destination', flat=True).distinct()
        extra_context['selected_bus_route'] = bus_route
        extra_context['selected_destination'] = destination
        extra_context['selected_class_no'] = class_no
        extra_context['selected_year'] = year
       

        return super().changelist_view(request, extra_context=extra_context)

    change_list_template = 'admin/transportdefaulter_change_list.html'

    def get_year(self):
        current_month = date.today().month
        return date.today().year if current_month >= 4 else date.today().year - 1
    
    # Modify get_search_results to return unique student IDs
    def get_search_results(self, request, queryset, search_term):
        year1 = request.GET.get('year', None)

        busno = request.GET.get('bus_route', None)
        destination = request.GET.get('destination', None)
        student_class = request.GET.get('class_no', None)
        if not busno and not year1 and not destination and not student_class:
            return queryset.none(), False
        year = int(year1) if year1 else self.get_year()
        current_date = datetime.now().date()
          # Get list of months for the year
        montharray = get_months_array(year)
        query = queryset.filter(
            year=year,
            bus_fees_paid__gt=0,
            student_id__bus_id__isnull=False
        )

        # print('query',query)

        route_subquery = RawSQL(
            "(SELECT route FROM busfees_master WHERE busfees_master.bus_id = student_master.bus_id)", []
        )

        destination_subquery = RawSQL(
            "(SELECT destination FROM busfees_master WHERE busfees_master.bus_id = student_master.bus_id)", []
        )

      
        query = query.annotate(
            route=RawSQL(
                "(SELECT route FROM busfees_master WHERE busfees_master.bus_id = student_master.bus_id)", []
            ),
            destination=RawSQL(
                "(SELECT destination FROM busfees_master WHERE busfees_master.bus_id = student_master.bus_id)", []
            ),
            # Avoid using GROUP_CONCAT in SQL here, as it might aggregate incorrectly
            months_paid=RawSQL(
                "(SELECT GROUP_CONCAT(TRIM(fees_period_month) ORDER BY CAST(fees_period_month AS UNSIGNED)) FROM student_fees WHERE student_fees.student_id = student_master.student_id AND student_fees.year = %s)",
                [year]
            )
        )


        if busno:
            query = query.filter(route=busno)

        if destination:
            query = query.filter(destination=destination)

        if student_class:
            query = query.filter(student_class=student_class)

        # Filter only students with unpaid months
        fees_records = []
        for data in query:
            months_paid = data.months_paid
            passedout_date = data.student_id.passedout_date

            if not months_paid:
                months_paid_list = []
            else:
                try:
                    months_paid_list = list(map(int, set(map(str.strip, months_paid.split(',')))))
                except ValueError:
                    months_paid_list = []

            unpaid_months = list(set(montharray) - set(months_paid_list))

            if unpaid_months and (not passedout_date or passedout_date >= current_date):
                fees_records.append(data.student_id.student_id)


        final_query = query.filter(student_id__in=fees_records).distinct()
        return final_query, False
    

admin.site.register(transport_defaulter, TransportDefaulterReportAdmin)


class chequedepositreportAdmin(admin.ModelAdmin):

    def has_add_permission(self, request):
        return False

    def has_change_permission(self, request, obj=None):
        return False

    def has_delete_permission(self, request, obj=None):
        return False
    change_list_template = "admin/chequedepositreport_change_list.html"
    
    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path('cheque-deposit/', self.admin_site.admin_view(self.changelist_view)),
            path('cheque-deposit/export/', self.admin_site.admin_view(self.export_cheque_deposits_to_excel), name="export_cheque_deposits_excel"),
        ]
        return custom_urls + urls
   

    def export_cheque_deposits_to_excel(seld, request):
        # datefrom = request.GET.get('datefrom')
        # dateto = request.GET.get('dateto')
        datefrom = request.session.get('datefrom',None)
        dateto = request.session.get('dateto',None)
        # Memory limit (optional in Python, memory-efficient libraries can be used)
        # Not necessary in Python but large datasets might require optimization

        bankarray = ['Bank', 'Cash', 'Online']
        datefrom1 = datetime.strptime(datefrom, "%Y-%m-%d").date()
        dateto1 = datetime.strptime(dateto, "%Y-%m-%d").date()
        date = datetime.now().year
        sessionstarts = f"{date}-04-01"

        # Create a new Excel workbook
        workbook = Workbook()
        sheet_index = 0

        for bank in bankarray:
            total_amount = 0
            if bank == 'Cash':
                query = student_fee.objects.filter(
                    payment_mode='Cash',
                    date_payment__range=[datefrom1, dateto1]
                ).select_related('student_id')
            elif bank == 'Online':
                query = student_fee.objects.filter(
                    payment_mode='online',
                    date_payment__range=[datefrom1, dateto1]
                ).select_related('student_id')
            else:
                query = student_fee.objects.filter(
                    payment_mode='cheque',
                    cheque_status='open',
                    date_payment__range=[datefrom1, dateto1]
                ).select_related('student_id')

            if query.exists():
                # Create a new sheet for each bank type
                if sheet_index == 0:
                    sheet = workbook.active
                    sheet.title = bank
                    # sheet = workbook.create_sheet(title=bank)
                else:
                    sheet = workbook.create_sheet(title=bank)
                sheet_index += 1

                # Set headers and school information
                sheet['C1'] = 'Shishu Niketan Public School, Sector 66, Mohali'
                sheet['C2'] = 'ICICI, Phase – 7, Mohali Account No. 632205010090'
                sheet['C3'] = 'Tel: 9815094449'
                sheet['E3'] = f'Date: {datetime.now().strftime("%d-%m-%Y")}'
                sheet['G3'] = f'Date of Entry: {datefrom1.strftime("%Y-%m-%d")}'

                # Set column headers
                headers = ['SrNo', 'Name', 'Admission Number(s)', 'Class', 'Section', 'Phone Number', 'Bank', 'Branch', 'Cheque No', 'Cheque Status', 'Amount']
                for col_num, header in enumerate(headers, 1):
                    sheet.cell(row=5, column=col_num).value = header

                # Fill in data
                row_num = 6
                for index, fee in enumerate(query, 1):
                    student = fee.student_id
                    phone_no = student.phone_no if student.phone_no else student.mobile_no

                    sheet.cell(row=row_num, column=1).value = index
                    sheet.cell(row=row_num, column=2).value = student.student_name
                    sheet.cell(row=row_num, column=3).value = student.addmission_no
                    sheet.cell(row=row_num, column=4).value = fee.student_class
                    sheet.cell(row=row_num, column=5).value = fee.student_section
                    sheet.cell(row=row_num, column=6).value = phone_no
                    sheet.cell(row=row_num, column=7).value = fee.bank_name
                    sheet.cell(row=row_num, column=8).value = fee.branch_name
                    sheet.cell(row=row_num, column=9).value = fee.cheq_no
                    sheet.cell(row=row_num, column=10).value = fee.cheque_status
                    sheet.cell(row=row_num, column=11).value = fee.amount_paid
                    total_amount += fee.amount_paid if fee.amount_paid else 0
                    row_num += 1

                # Add total row
                sheet.cell(row=row_num, column=1).value = 'Total'
                sheet.cell(row=row_num, column=11).value = total_amount
                print('bank',bank)
                print('sheet_index',sheet_index)

        # Create a response for the Excel file download
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="Cheque_Deposit_Status_Report.xlsx"'

        # Save the workbook to the response
        workbook.save(response)
        return response

    def changelist_view(self, request, extra_context=None):
        extra_context = extra_context or {}
        cheque_details = []
        total_amount = 0

        # Retrieve date filters from the POST data
        if request.method == 'POST':
            datefrom = request.POST.get('datefrom')
            dateto = request.POST.get('dateto')
            print('datefrom',datefrom)
            print('dateto',dateto)
            # Validate and parse dates
            try:
                datefrom = datetime.strptime(datefrom, "%Y-%m-%d").date() if datefrom else None
                dateto = datetime.strptime(dateto, "%Y-%m-%d").date() if dateto else None
            except ValueError:
                datefrom, dateto = None, None

            # Set defaults if dates are missing
            if not datefrom:
                datefrom = datetime.now().replace(month=4, day=1).date()  # Start of fiscal year
            if not dateto:
                dateto = datetime.now().date()
                

            print(f'Date From: {datefrom}, Date To: {dateto}')

            # Construct and execute the SQL query
            query = f"""
                SELECT sf.cheq_no, sf.bank_name, sf.branch_name, sf.cheque_status, 
                    SUM(sf.amount_paid) as amount_paid, 
                    GROUP_CONCAT(DISTINCT sm.student_name) as student_name, 
                    sm.mobile_no,
                    GROUP_CONCAT(DISTINCT sm.addmission_no) as addmission_no, 
                    GROUP_CONCAT(DISTINCT sf.student_class) as student_class, 
                    GROUP_CONCAT(DISTINCT sf.student_section) as student_section 
                FROM student_fees sf 
                LEFT JOIN student_master sm ON sm.student_id = sf.student_id 
                WHERE sf.payment_mode='cheque' 
                AND sf.cheque_status='open' 
                AND sf.date_payment BETWEEN '{datefrom}' AND '{dateto}'
                GROUP BY sf.bank_name, sf.branch_name, sf.cheq_no
            """
            params = [datefrom, dateto]

            print(f'Query: {query}')
            print(f'Params: {params}')

            # Execute the query
            with connection.cursor() as cursor:
                # cursor.execute(query, params)
                cursor.execute(query)
                query_results = cursor.fetchall()

            # print(f'Query Results: {query_results}')

            # Process the query results
            for row in query_results:
                cheque_details.append({
                    'student_name': row[5],
                    'admission_no': row[7],
                    'student_class': row[8],
                    'student_section': row[9],
                    'mobile_no': row[6],  # Corrected the mobile number index
                    'bank_name': row[1],
                    'branch_name': row[2],
                    'cheq_no': row[0],
                    'cheque_status': row[3],
                    'amount_paid': row[4],
                })
                total_amount += row[4]  # Sum up the total amount paid

            # print('cheque_details',cheque_details)
            # Store the results in the session
            datefrom_str = datefrom.isoformat()  # Converts to 'YYYY-MM-DD' string
            dateto_str = dateto.isoformat()  
            request.session['datefrom'] = datefrom_str
            request.session['dateto'] = dateto_str

        # Add the cheque details and total amount to the context
        extra_context['cheque_details'] = cheque_details
        extra_context['total_amount'] = total_amount
        extra_context['today_date'] = datetime.now().strftime("%Y-%m-%d")

        return super().changelist_view(request, extra_context=extra_context)

admin.site.register(cheque_deposit, chequedepositreportAdmin)





class CollectionReportForm(forms.Form):
    FEE_TYPE_CHOICES = [
        ('advancedfees', 'Advanced fees'),
        ('totalfees', 'Total fees'),
        ('customfees', 'LastDay fees'),
    ]

    PAYMENT_TYPE_CHOICES = [
        ('', 'Payment'),
        ('Online', 'Online'),
        ('Cash', 'Cash'),
        ('Cheque', 'Cheque'),
    ]

    DEVICE_TYPE_CHOICES = [
        ('', 'Device'),
        ('WEB', 'WEB'),
        ('MOBILE', 'MOBILE'),
    ]

    feestype = forms.ChoiceField(
        choices=FEE_TYPE_CHOICES, 
        widget=forms.RadioSelect,
        label='Fee Type'
    )
    
    paymentType = forms.ChoiceField(
        choices=PAYMENT_TYPE_CHOICES,
        required=False,
        label='Payment Type',
        widget=forms.Select(attrs={'id': 'paymentType'})
    )
    
    deviceType = forms.ChoiceField(
        choices=DEVICE_TYPE_CHOICES,
        required=False,
        label='Device Type',
        widget=forms.Select(attrs={'id': 'deviceType'})
    )
    
    date_from = forms.DateField(
        required=False,
        label='Date From',
        widget=forms.DateInput(attrs={'id': 'id_date_from', 'type': 'date'})
    )
    
    date_to = forms.DateField(
        required=False,
        label='Date To',
        widget=forms.DateInput(attrs={'id': 'id_date_to', 'type': 'date'})
    )

    filter_button = forms.CharField(
        widget=forms.TextInput(attrs={'type': 'submit', 'value': 'Filter'}),
        label='',
        required=False
    )

class CollectionReportAdmin(admin.ModelAdmin):
   
    change_list_template = "admin/collection_report/change_list.html"  # Specify your custom template

    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path('collection_report/', self.admin_site.admin_view(self.changelist_view)),
            path('collection_report/export/', self.admin_site.admin_view(self.export_defaulters_to_excel), name="export_collection_report_csv")
           
        ]
        return custom_urls + urls

    def has_add_permission(self, request):
        return False

    def has_change_permission(self, request, obj=None):
        return False

    def has_delete_permission(self, request, obj=None):
        return False

    def export_defaulters_to_excel(self, request):
        # Retrieve data from session, stored as a JSON string
        collection_list_str = request.session.get('collection_list', '')

        if not collection_list_str:
            # Handle the case where there is no data to export
            messages.error(request, "No data available to export.")
            return redirect('admin:collection_report')

        # Convert the string to a list of dictionaries
        try:
            collection_list = json.loads(collection_list_str)
            print(f"collection_list content after loading JSON: {collection_list}")
            print(f"collection_list type after loading JSON: {type(collection_list)}")
        except json.JSONDecodeError as e:
            messages.error(request, f"Error decoding JSON: {e}")
            return redirect('admin:collection_report')

        # Define the column headers
        headers = [
            "Admission No", "Student Name", "Class", "Section", "Fees for Months", 
            "Fees Period Months", "Annual Fees Paid", "Tuition Fees Paid", "Concession Applied", 
            "Net Tuition Fees", "Funds Fees Paid", "Sports Fees Paid", "Activity Fees", 
            "Admission Fees Paid", "Security Fees Paid", "Late Fees Paid", "Miscellaneous Fees", 
            "Bus Fees Paid", "Date of Payment", "Payment Mode", "Cheque No.", "Bank Name", 
            "Concession Type", "Total Amount", "Amount Paid", "Cheque Status", 
            "Realized Date", "Remarks", "Request Source"
        ]

        try:
            # Ensure that collection_list is a list of dictionaries
            if isinstance(collection_list, list) and all(isinstance(row, dict) for row in collection_list):
                # Create a DataFrame from the list of dictionaries
                df = pd.DataFrame(collection_list)

                # Add the calculated columns (Net Tuition Fees and Miscellaneous Fees)
                df['Net Tuition Fees'] = df['tuition_fees_paid'] - df['concession_applied']
                df['Miscellaneous Fees'] = (
                    df['funds_fees_paid'] + df['sports_fees_paid'] + df['activity_fees'] + 
                    df['late_fees_paid']
                )
                
                # Ensure the DataFrame matches the required headers
                df = df.reindex(columns=[
                    'addmission_no', 'student_name', 'class_no', 'section', 'fees_for_months', 
                    'fees_period_month', 'annual_fees_paid', 'tuition_fees_paid', 'concession_applied',
                    'Net Tuition Fees', 'funds_fees_paid', 'sports_fees_paid', 'activity_fees',
                    'admission_fees_paid', 'security_paid', 'late_fees_paid', 'Miscellaneous Fees', 
                    'bus_fees_paid', 'date_payment', 'payment_mode', 'cheq_no', 'bank_name', 
                    'concession_type', 'Total_amount', 'amount_paid', 'cheque_status', 
                    'realized_date', 'remarks', 'request_source'
                ])
                
                # Rename the columns to match the headers you provided
                df.columns = headers
            else:
                print(f"Unexpected collection_list format: {collection_list}")
                raise ValueError("Unsupported format in collection_list. Must be list of dictionaries.")
        except Exception as e:
            # Handle any errors during DataFrame creation or renaming
            messages.error(request, f"Error during data export: {e}")
            return redirect('admin:collection_report')

        # Create a BytesIO stream to hold the Excel data
        excel_file = BytesIO()
        try:
            with pd.ExcelWriter(excel_file, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, sheet_name='Defaulters')

            # Rewind the buffer
            excel_file.seek(0)

            # Create the HTTP response to download the file
            response = HttpResponse(
                excel_file.read(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = 'attachment; filename=collection_report.xlsx'
            return response
        except Exception as e:
            # Handle errors during Excel file generation
            messages.error(request, f"Error generating Excel file: {e}")
            return redirect('admin:collection_report')

    def changelist_view(self, request, extra_context=None):

        form = CollectionReportForm(request.POST or None)

        def date_to_string(obj):
            if isinstance(obj, date):
                return obj.strftime('%Y-%m-%d')  # You can choose another format if needed
            raise TypeError(f'Object of type {obj.__class__.__name__} is not JSON serializable')

        # results = None  # Initialize results to None
        results = []  # Change this to a list to hold dictionaries

        if request.method == 'POST' and form.is_valid():
            # Get the cleaned form data
            feestype = form.cleaned_data.get('feestype')
            paymentType = form.cleaned_data.get('paymentType')
            deviceType = form.cleaned_data.get('deviceType')
            date_from = form.cleaned_data.get('date_from')
            date_to = form.cleaned_data.get('date_to')

            # Initialize query variables
            feedetails = ""
            amount = 0
            totsecfees = 0  # Initialize to 0
            # Initialize other totals...

            current_year = datetime.now().year
            session_starts = f"{current_year}-04-01"

            year = current_year - 1
            toyear = year + 1
            from_date = f"{year}-04-01"
            to_date = f"{toyear}-03-31"

            # Check if date_from and date_to are strings before parsing
            datefrom1 = parse_date(date_from) if isinstance(date_from, str) else None
            dateto1 = parse_date(date_to) if isinstance(date_to, str) else None

            query = ""

            if feestype == 'advancedfees':
                query = f"""
                    SELECT sf.*, sf.student_class as class_no, sf.student_section as section, 
                        sm.addmission_no, sm.student_name, con.concession_type, sf.request_source
                    FROM student_fees sf
                    INNER JOIN student_master sm ON sm.student_id = sf.student_id 
                    LEFT JOIN concession_master con ON con.concession_id = sf.concession_type_id
                    WHERE sf.year = '{current_year}' 
                    AND sf.date_payment BETWEEN '{from_date}' AND '{to_date}'
                """

                if paymentType:
                    query += f" AND sf.payment_mode = '{paymentType}'"
                if deviceType:
                    query += f" AND sf.request_source = '{deviceType}'"

                query += " ORDER BY sf.date_payment DESC, sm.addmission_no;"

            elif feestype == 'customfees':
                query = f"""
                    SELECT sf.*, cls.*, sm.addmission_no, sm.student_name, con.concession_type, sf.request_source 
                    FROM student_fees sf 
                    LEFT JOIN student_master sm ON sm.student_id = sf.student_id 
                    LEFT JOIN concession_master con ON con.concession_id = sf.concession_type_id 
                    LEFT OUTER JOIN (
                        SELECT * FROM student_classes 
                        WHERE (student_id, started_on) IN (
                            SELECT student_id, MAX(started_on) 
                            FROM student_classes 
                            GROUP BY student_id
                        )
                    ) AS cls ON sm.student_id = cls.student_id
                """

                if datefrom1 and dateto1:
                    query += f" WHERE sf.date_payment BETWEEN '{datefrom1}' AND '{dateto1}'"
                else:
                    query += " WHERE sf.added_at >= CURDATE() - INTERVAL 1 DAY AND sf.added_at < CURDATE()"

                if deviceType:
                    query += f" AND sf.request_source = '{deviceType}'"
                if paymentType:
                    query += f" AND sf.payment_mode = '{paymentType}'"

                query += " ORDER BY sf.added_at DESC, sm.addmission_no;"

            else:
                query = f"""
                    SELECT a.*, cls.*, sm.addmission_no, sm.student_name, con.concession_type, a.request_source 
                    FROM student_fees a 
                    LEFT JOIN student_master sm ON sm.student_id = a.student_id 
                    LEFT JOIN concession_master con ON con.concession_id = a.concession_type_id 
                    LEFT OUTER JOIN student_classes cls ON sm.student_id = cls.student_id 
                    WHERE a.date_payment BETWEEN '{datefrom1}' AND '{dateto1}'
                    AND cls.student_class_id = (
                        SELECT MAX(student_class_id) 
                        FROM student_classes 
                        WHERE student_id = sm.student_id
                    )
                """

                if paymentType:
                    query += f" AND a.payment_mode = '{paymentType}'"
                if deviceType:
                    query += f" AND a.request_source = '{deviceType}'"

                query += " ORDER BY a.date_payment DESC, sm.addmission_no;"
            
            amount = 0
            totsecfees = 0  # Initialize to 0

            # Execute the query and process the results
            with connection.cursor() as cursor:
                cursor.execute(query)
                result = cursor.fetchall()

                # Create a mapping of column names to their indices
                column_names = [col[0] for col in cursor.description]

                for row in result:
                    fee_record = {
                        'addmission_no': row[column_names.index('addmission_no')],
                        'student_name': row[column_names.index('student_name')],
                        'class_no': row[column_names.index('class_no')],
                        'section': row[column_names.index('section')],
                        'fees_for_months': row[column_names.index('fees_for_months')],
                        'fees_period_month': row[column_names.index('fees_period_month')],
                        'annual_fees_paid': row[column_names.index('annual_fees_paid')],
                        'tuition_fees_paid': row[column_names.index('tuition_fees_paid')],
                        'concession_applied': row[column_names.index('concession_applied')],
                        # 'net_tuition_fees': row[column_names.index('net_tuition_fees')],
                        'funds_fees_paid': row[column_names.index('funds_fees_paid')],
                        'sports_fees_paid': row[column_names.index('sports_fees_paid')],
                        'activity_fees': row[column_names.index('activity_fees')],
                        'admission_fees_paid': row[column_names.index('admission_fees_paid')],
                        'security_paid': row[column_names.index('security_paid')],
                        'late_fees_paid': row[column_names.index('late_fees_paid')],
                        # 'miscellaneous_fees': row[column_names.index('miscellaneous_fees')],
                        'bus_fees_paid': row[column_names.index('bus_fees_paid')],
                        'date_payment': row[column_names.index('date_payment')],
                        'payment_mode': row[column_names.index('payment_mode')],
                        'cheq_no': row[column_names.index('cheq_no')],
                        'bank_name': row[column_names.index('bank_name')],
                        'concession_type': row[column_names.index('concession_type')],
                        'Total_amount': row[column_names.index('Total_amount')],
                        'amount_paid': row[column_names.index('amount_paid')],
                        'cheque_status': row[column_names.index('cheque_status')],
                        'realized_date': row[column_names.index('realized_date')],
                        'remarks': row[column_names.index('remarks')],
                        'request_source': row[column_names.index('request_source')]
                    }

                    results.append(fee_record)  # Add the fee record to the list
                    amount += row[column_names.index('amount_paid')] or 0  # Handle None for amount as well

        # Ensure any date objects in the result are converted to strings
        for result in results:
            if 'date_payment' in result and isinstance(result['date_payment'], date):
                result['date_payment'] = result['date_payment'].strftime('%Y-%m-%d')
            if 'realized_date' in result and isinstance(result['realized_date'], date):
                result['realized_date'] = result['realized_date'].strftime('%Y-%m-%d')

        # Now save the results to the session using json.dumps with the date_to_string function
        request.session['collection_list'] = json.dumps(results, default=date_to_string)

        # Prepare context to pass to the template
        context = extra_context or {}
        context.update({
            'form': form,
            'results': results,  # Pass the list of dictionaries to the template
        })

        # Render the custom template with the form and results
        return super().changelist_view(request, extra_context=context)

admin.site.register(collection_report, CollectionReportAdmin)

class DefaultersReportForm(forms.Form):
    CLASS_CHOICES = [
        ('', 'Select Class'),
        ('nursery', 'Nursery'),
        ('LKG', 'LKG'),
        ('UKG', 'UKG'),
        ('1', '1st Grade'),
        ('2', '2nd Grade'),
        # Add other classes as needed
    ]

    current_year = datetime.now().year
    YEAR_CHOICES = [(str(year), str(year)) for year in range(current_year, current_year - 6, -1)]

    student_class = forms.ChoiceField(choices=CLASS_CHOICES, required=True, label="Class", widget=forms.Select(attrs={'id': 'id_student_class'}))
    year = forms.ChoiceField(choices=YEAR_CHOICES, required=True, label="Year",widget=forms.Select(attrs={'id': 'id_year'}))

    # Search button with jQuery to trigger a student search
    search_button = forms.CharField(
        widget=forms.TextInput(attrs={'type': 'submit', 'value': 'Search'}),
        label='',
        required=False
    )

class ActivityFeesDefaulterAdmin(admin.ModelAdmin):

    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path('activity-fees-defaulters/', self.admin_site.admin_view(self.changelist_view),name='activity_fees_defaulter_changelist'),
            path('activity-fees-defaulters/export/', self.admin_site.admin_view(self.export_defaulters_to_excel), name="export_defaulters_csv")
        ]
        return custom_urls + urls

    def has_add_permission(self, request):
        return False

    def has_change_permission(self, request, obj=None):
        return False

    def has_delete_permission(self, request, obj=None):
        return False

    def get_months_array(self, year):
        # Convert year to an integer for arithmetic operations
        year = int(year)

        # Define the start and end of the financial year
        datefrom = f"{year}0401"  # First date of the financial year (April 1st)
        dateto = f"{year + 1}0331"  # Last date of the financial year (March 31st of the next year)

        # Convert to actual date format (Y-m-d)
        datefrom1 = datetime.strptime(datefrom, "%Y%m%d").date()
        dateto1 = datetime.strptime(dateto, "%Y%m%d").date()

        # Get the current date
        currentdate = datetime.now().date()

        # If the financial year is not over yet, set `dateto1` to today
        if currentdate < dateto1:
            dateto1 = currentdate

        # Initialize the month array
        month_array = []
        
        # Create a timestamp starting from `datefrom1`
        time = datefrom1
        to_month = dateto1.month

        # Iterate for a maximum of 12 months
        for i in range(12):
            # Get the current month
            cur_month = time.month
            
            # Stop if the last month (current month) is reached
            if cur_month == to_month:
                break

            # Add the month number to the array, stripping leading zeroes
            month_array.append(str(cur_month).lstrip('0'))
            
            # Move to the next month
            time = time + timedelta(days=31)
            time = time.replace(day=1)  # Ensure to set the day to the 1st of the next month

        # If all months have passed, return a full array from 1 to 12
        if not month_array:
            month_array = [str(i) for i in range(1, 13)]

        # Sort the month array to ensure the months are in ascending order
        month_array.sort(key=int)

        return month_array


    def export_defaulters_to_excel(self, request):
        # Retrieve data to export, typically from a session or request
        defaulters_list = request.session.get('defaulters_list', [])
        
        if not defaulters_list:
            # Handle the case where there is no data to export
            messages.error(request, "No data available to export.")
            return redirect('admin:activity_fees_defaulter_changelist')

        # Define the column headers
        headers = ['Admission No', 'Student Name', 'Class', 'Section', 'Unpaid Months']

        # Convert defaulters_list to DataFrame for export
        df = pd.DataFrame(defaulters_list)

        # Rename the columns to match the desired headers
        df.columns = headers

        # Create a BytesIO stream to hold the Excel data
        excel_file = BytesIO()
        with pd.ExcelWriter(excel_file, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Defaulters')

        # Rewind the buffer
        excel_file.seek(0)

        # Create the HTTP response to download the file
        response = HttpResponse(
            excel_file.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=defaulters_report.xlsx'
        
        return response

    def changelist_view(self, request, extra_context=None):

        extra_context = extra_context or {}
        form = DefaultersReportForm(request.POST or None)  # Initialize form with POST data

        results = []
        if request.method == 'POST' and form.is_valid():
            selected_class = form.cleaned_data.get('student_class')
            selected_year = form.cleaned_data.get('year')

            # print(f"selected_class-----{selected_class}")
            # print(f"selected_year-----{selected_year}")

            if selected_year is None:
                current_year = date.today().year
                selected_year = current_year - 1 if date.today().month < 4 else current_year

            # Prepare data for queries
            current_date = date.today()
            month_array = self.get_months_array(selected_year)
            months_paid_str = ",".join(month_array)

            # print(f"current_date-----{current_date}")
            # print(f"month_array-----{month_array}")
            # print(f"months_paid_str-----{months_paid_str}")

            # class_string = "'nursery', 'pre-nursery', 'kg', '1', '2'"  # Add other classes as needed
            class_string = 'nursery,pre-nursery,kg,1,2'

            # Query to fetch students who have paid but not fully
            query = f"""
            SELECT student_fees.student_id, 
                GROUP_CONCAT(TRIM(student_fees.fees_period_month) ORDER BY CAST(student_fees.fees_period_month AS UNSIGNED)) AS months_paid,
                student_fees.student_class AS class_no,
                student_master.addmission_no, 
                student_master.admission_date,
                student_master.passedout_date, 
                student_master.student_name, 
                student_classes.section
            FROM student_fees
            LEFT JOIN student_master ON student_fees.student_id = student_master.student_id
            LEFT JOIN student_classes ON student_fees.student_id = student_classes.student_id 
                                    AND student_classes.class_no = student_fees.student_class
            WHERE student_fees.year = %s 
            AND COALESCE(student_fees.activity_fees, 0) > 0
            AND FIND_IN_SET(student_fees.student_class, %s)
            """

            if selected_class:
                query += f" AND student_fees.student_class = '{selected_class}' "

            query += f"""
            GROUP BY student_fees.student_id
            HAVING months_paid != '{months_paid_str}'
            """

            # Query to fetch students who haven't paid at all
            query_union = f"""
            SELECT student_master.student_id, 
                '' AS months_paid, 
                student_classes.class_no, 
                student_master.addmission_no,
                student_master.admission_date, 
                student_master.passedout_date, 
                student_master.student_name, 
                student_classes.section
            FROM student_master
            INNER JOIN student_classes ON student_master.student_id = student_classes.student_id
            WHERE student_master.student_id NOT IN (
                SELECT student_id FROM student_fees WHERE year = %s
            )
            AND FIND_IN_SET(student_classes.class_no, %s)
            """

            if selected_class:
                query_union += f" AND student_classes.class_no = '{selected_class}' "

            query_union += " GROUP BY student_classes.student_id HAVING MAX(student_classes.class_no)"

            # print(f"query-----{query}")
            # print(f"query_union-----{query_union}")

            # Combine both queries
            final_query = f"SELECT * FROM ({query} UNION {query_union}) AS combined ORDER BY class_no, section, addmission_no"

            

            # Execute the query using Django's raw SQL execution
            with connection.cursor() as cursor:
                # Manually format the query with actual parameter values for logging
                formatted_query = final_query % (selected_year, class_string, selected_year, class_string)

                # Print the formatted query to the console
                # print(f"final_query-----{formatted_query}")

                # Execute the query using the parameters
                cursor.execute(final_query, [selected_year, class_string, selected_year, class_string])

                # Fetch and print the results
                query_results = cursor.fetchall()
                # print(f"query_results-----{query_results}")

            # with connection.cursor() as cursor:
            #     cursor.execute(final_query, [selected_year, class_string, selected_year, class_string])
            #     print(f"final_query-----{final_query}")
            #     query_results = cursor.fetchall()
            #     print(f"query_results-----{query_results}")

            # Process results to identify unpaid months and defaulters
            defaulters_list = []
            for data in query_results:
                months_paid = data[1]
                passedout_date = data[5]
                months_paid_cleaned = [month.strip() for month in months_paid.split(',') if month]
                months_unpaid = sorted(list(set(month_array) - set(months_paid_cleaned)), key=lambda x: int(x))
                # months_unpaid = list(set(month_array) - set(months_paid_cleaned))

                # Only include students who have unpaid months and haven't passed out
                # if months_unpaid and (not passedout_date or passedout_date >= str(current_date)):
                if months_unpaid and (not passedout_date or passedout_date >= current_date):
                    defaulter_record = {
                        'addmission_no': data[3],
                        'student_name': data[6],
                        'class_no': data[2],
                        'section': data[7],
                        'unpaid_months': ",".join(months_unpaid)
                    }
                    defaulters_list.append(defaulter_record)

            # Save results to session for later export
            request.session['defaulters_list'] = defaulters_list

            results = defaulters_list

        # Add form and results to extra_context
        extra_context['form'] = form
        extra_context['results'] = results

        return super().changelist_view(request, extra_context=extra_context)

    change_list_template = 'admin/activity_fees_defaulter/change_list.html'

admin.site.register(activity_fees_defaulter, ActivityFeesDefaulterAdmin)



class ExpectedFeeRegisterForm(forms.Form):
    
    current_year = datetime.now().year
    # Prepend a "Select Year" option to the year choices
    YEAR_CHOICES = [('', 'Select Year')] + [(str(year), str(year)) for year in range(current_year, current_year - 11, -1)]

    year = forms.ChoiceField(choices=YEAR_CHOICES, required=True, label="Year",widget=forms.Select(attrs={'id': 'id_year'}))

    # Search button with jQuery to trigger a student search
    search_button = forms.CharField(
        widget=forms.TextInput(attrs={'type': 'submit', 'value': 'Export to excel'}),
        label='',
        required=False
    )

class ExpectedFeeRegisterAdmin(admin.ModelAdmin):
    change_list_template = 'admin/expected_fee_register/change_list.html'

    def get_form(self, request):
        """ Return the form that will be rendered in the changelist view. """
        return ExpectedFeeRegisterForm()

    def changelist_view(self, request, extra_context=None):
        form = self.get_form(request)

        # Check if form was submitted and is valid
        if request.method == 'POST':
            form = ExpectedFeeRegisterForm(request.POST)
            if form.is_valid():
                # Fetch the selected year from the form
                year = form.cleaned_data['year']
                
                # Instead of showing the search results, generate the Excel file
                return self.action_expected_fee_excel(year)

        # Display form as normal if no form submission occurred
        extra_context = extra_context or {}
        extra_context['form'] = form

        return super().changelist_view(request, extra_context=extra_context)

    
    def action_expected_fee_excel(request, year=None):
        # Initialize Excel file
        wb = Workbook()
        ws = wb.active
        ws.title = "Expected Fees Register"

        # Excel metadata
        wb.create_sheet(title=f"summary{year}")
        
        # Define year range
        to_year = int(year) + 1
        from_date = f"{year}-04-01"
        to_date = f"{to_year}-03-31"
        
        # First Query - Summary of Fees
        summary_sql = """
            SELECT sc.class_no, sc.section, COUNT(DISTINCT sc.student_id) as count_students, 
                SUM(fm.annual_fees), SUM(fm.tuition_fees), SUM(fm.funds_fees),
                SUM(fm.sports_fees), SUM(fm.activity_fees), SUM(fm.admission_fees), 
                SUM(fm.security_fees), SUM(fm.dayboarding_fees),
                (SUM(fm.annual_fees) + SUM(fm.tuition_fees) + SUM(fm.funds_fees) + 
                SUM(fm.sports_fees) + SUM(fm.activity_fees) + SUM(fm.admission_fees) +
                SUM(fm.security_fees) + SUM(fm.dayboarding_fees)) AS total
            FROM student_classes sc
            LEFT JOIN fees_master fm ON fm.class_no = sc.class_no
            WHERE sc.started_on >= %s AND (sc.ended_on IS NULL OR sc.ended_on <= %s)
            AND fm.valid_from >= %s AND (fm.valid_to IS NULL OR fm.valid_to <= %s)
            GROUP BY sc.class_no, sc.section
            ORDER BY sc.class_no, sc.section ASC
        """
        
        # Execute SQL and fetch data
        with connection.cursor() as cursor:
            cursor.execute(summary_sql, [from_date, to_date, from_date, to_date])
            summary_data = cursor.fetchall()
        
        # Excel Header
        headers = ['Class', 'Section', 'Total Students', 'Annual', 'Tuition', 'Funds', 
                'Sports', 'Activity', 'Admission', 'Security', 'Dayboarding', 'Total']
        ws.append(headers)
        
        # Write data to Excel
        total_students = total_annual_fees = total_tuition_fees = total_funds_fees = 0
        total_sports_fees = total_activity_fees = total_admission_fees = total_security_fees = 0
        total_dayboarding_fees = total_fees = 0
        
        for row in summary_data:
            class_no, section, count_students, annual_fees, tuition_fees, funds_fees, \
            sports_fees, activity_fees, admission_fees, security_fees, dayboarding_fees, total = row
            
            annual_fees *= 12
            tuition_fees *= 12
            funds_fees *= 12
            sports_fees *= 12
            activity_fees *= 12
            admission_fees *= 12
            security_fees *= 12
            dayboarding_fees *= 12
            total = annual_fees + tuition_fees + funds_fees + sports_fees + \
                    activity_fees + admission_fees + security_fees + dayboarding_fees
            
            # Append row data to Excel
            ws.append([class_no, section, count_students, annual_fees, tuition_fees, 
                    funds_fees, sports_fees, activity_fees, admission_fees, 
                    security_fees, dayboarding_fees, total])
            
            # Accumulate totals
            total_students += count_students
            total_annual_fees += annual_fees
            total_tuition_fees += tuition_fees
            total_funds_fees += funds_fees
            total_sports_fees += sports_fees
            total_activity_fees += activity_fees
            total_admission_fees += admission_fees
            total_security_fees += security_fees
            total_dayboarding_fees += dayboarding_fees
            total_fees += total
        
        # Add Totals row
        ws.append(['Total', '', total_students, total_annual_fees, total_tuition_fees, 
                total_funds_fees, total_sports_fees, total_activity_fees, total_admission_fees,
                total_security_fees, total_dayboarding_fees, total_fees])

        # Set HTTP response
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="Expected_Fees_{year}.xlsx"'
        
        # Save workbook to response
        wb.save(response)
        return response


admin.site.register(expected_fee_register, ExpectedFeeRegisterAdmin)